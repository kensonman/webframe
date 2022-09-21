# -*- coding: utf-8 -*-
#
# File: src/webframe/management/commands/pref.py
# Date: 2020-04-22 21:35
# Author: Kenson Man <kenson@kenson.idv.hk>
# Desc: Import / Create / Update / Delete preference
#
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q
from pathlib import Path
from webframe.functions import TRUE_VALUES, LogMessage as lm, getTime
from webframe.models import Preference, AbstractPreference
from uuid import UUID
import logging, os, glob, sys, re


logger=logging.getLogger('webframe.commands.prefs')

class Command(BaseCommand):
   help     = '''Mainpulate the preference in database. Including insert/update/delete/view/import/gensecret/gendoc; Importing support csv|xlsx file.'''

   def __getIndent__(self, indent=0, ch=' '):
      return ch*indent

   def create_parser(self, cmdName, subcommand, **kwargs):
      parser=super().create_parser(cmdName, subcommand, **kwargs)
      parser.epilog='''Example:\r\n
\tpref import path_to_prefs   #Import a folder or a csv/xlsx file\r\n
\tpref set ABC --value="def"  #Set the preference "ABC" to value "def"\r\n
\tpref gensecret              #Generate the encryption secret; PLEASE backup in secure way.\r\n
\tpref gendoc prefsDoc.html     #Generate the documentation and save as as output.html
'''
      return parser

   def add_arguments(self, parser):
      #Default Value
      pattern='Pref({pref.id}:{pref.name}): {pref.value}'
      action='show'
      max=256
      wildcard='*'
      tmpl='webframe/prefsDoc.html'

      #Adding arguments
      parser.add_argument('action', type=str, help='The action to be taken. One of import/export/show/set/delete/gensecret/gendoc; Default is {0}'.format(action), default=action)
      parser.add_argument('name', type=str, nargs='?', help='[import/export/show/set/delete/gendoc]; The name of the preference or path of importing/exporting file (csv|xlsx);')
      parser.add_argument('--file', dest='file', type=str, help='[import/export/gendoc]; The file path for import/export/output.')
      parser.add_argument('--value', dest='value', type=str, help='[set/delete]; The value of the preference;', default=None)
      parser.add_argument('--owner', dest='owner', type=str, help='[set/delete]; The owner of the preference; Optional;', default=None)
      parser.add_argument('--noowner', dest='noowner', action='store_true', help='[show/set/delete]; The target preference has no owner; Optional; Default False')
      parser.add_argument('--parent', dest='parent', type=str, help='[show/set/delete]; The parent\'s name of the preference. Optional;', default=None)
      parser.add_argument('--noparent', dest='noparent', action='store_true', help='[show/set/delete]; The target preference has no parent; Optional; Default False')
      parser.add_argument('--pattern', dest='pattern', type=str, help='[show]; The output pattern. {0}'.format(pattern), default=pattern)
      parser.add_argument('--max', dest='max', type=int, help='[show]; The maximum number of preference to show. Default is {0}'.format(max), default=max)
      parser.add_argument('--wildcard', dest='wildcard', type=str, help='[show]; Specify the wildcard; Default is {0}'.format(wildcard), default=wildcard)

      #Importing 
      parser.add_argument('--sep', dest='separator', type=str, default=',', help='[import]; The separator when CSV importing; Default \",\"')
      parser.add_argument('--encoding', dest='encoding', type=str, default='utf-8', help='[import]; The encoding when CSV importing; Default \"utf-8\"')
      parser.add_argument('--quotechar', dest='quotechar', type=str, default='\"', help='[import]; The quote-char when CSV importing; Default double quote: \"')
      parser.add_argument('--filepath', dest='filepath', action='store_true', help='[import]; Import the file-path in preferences; Default False')
      parser.add_argument('--force', '-f ', dest='force', action='store_true', help='[import]; Force the import', default=False)

      #Generate Doc
      parser.add_argument('--tmpl', dest='tmpl', type=str, help="[gendoc]; The template name when generating document; Default: {0}".format(tmpl), default=tmpl)

   def __get_owner__(self, owner=None):
      if not owner: return None
      logger.debug('Getting owner by: "%s"', owner)
      owner=owner if owner else self.kwargs['owner']
      return get_user_model().objects.get(username=owner) if owner else None

   def __get_parent__(self, parent=None):
      parent=parent if parent else self.kwargs['parent']
      if parent:
         try:
            #Get parent by uuid
            return Preference.objects.get(id=parent)
         except:
            try:
               #Get parent by name
               return Preference.objects.get(name=parent)
            except:
               pass
      return None

   def __get_pref__(self, **kwargs):
      owner=kwargs['owner'] if 'owner' in kwargs else self.__get_owner__()
      parent=kwargs['parent'] if 'parent' in kwargs else self.__get_parent__()
      name=kwargs['name'] if 'name' in kwargs else self.kwargs['name']
      lang=kwargs['lang'] if 'lang' in kwargs else None
      if self.kwargs['filepath']: name=os.path.basename(name)

      if self.kwargs['parent'] and parent==None:
         raise Preference.DoesNotExist('Parent Preference not found: {0}'.format(self.kwargs['parent']))

      rst=Preference.objects.all()
      if name and name!='*':
         rst=rst.filter(name=name)
      if owner:
         rst=rst.filter(owner=owner)
      elif self.kwargs['noowner']:
         rst=rst.filter(owner__isnull=True)

      if parent:
         rst=rst.filter(parent=parent)
      elif self.kwargs['noparent']:
         rst=rst.filter(parent__isnull=True)

      if self.kwargs['filepath']:
         rst=rst.filter(tipe=AbstractPreference.TYPE_FILEPATH)

      rst=rst.order_by('owner', 'parent', 'sequence', 'name')

      return rst

   def __get_name__( self, name ):
      '''
      Get the name and sequence according to the name.
      @param name The string including the sequence and name. For example, '01.Target' will return a tuple (1, 'Target')
      @return A tuple including the sequence and the name
      '''
      p=re.search(r'^\d+\.', name)
      if p:
         s=p.group(0)
         return name[len(s):].strip(), int(name[0:len(s)-1])
      return (name, sys.maxsize if hasattr(sys, 'maxsize') else sys.maxint) #Default append

   def output( self, pref, pattern=None ):
      pattern=pattern if pattern else self.kwargs['pattern']
      print(pattern.format(pref=pref))
      pattern='   {0}'.format(pattern)
      for ch in pref.childs:
         self.output(ch, pattern)

   def handle(self, *args, **kwargs):
      verbosity=int(kwargs['verbosity'])
      if verbosity==3:
         logger.setLevel(logging.DEBUG)
      elif verbosity==2:
         logger.setLevel(logging.INFO)
      elif verbosity==1:
         logger.setLevel(logging.WARNING)
      else:
         logger.setLevel(logging.ERROR)
      self.kwargs=kwargs

      action=kwargs['action']
      if action=='import':
         self.imp()
      elif action=='create': #for backward compatibility
         self.set()
      elif action=='update': #for backward compatibility
         self.set()
      elif action=='set':
         self.set()
      elif action=='delete':
         self.delete()
      elif action=='show':
         self.show()
      elif action=='gensecret':
         self.gensecret()
      elif action=='gendoc':
         self.gendoc()
      elif action=='export':
         self.expCsv()
      else:
         logger.warning('Unknown action: {0}'.format(action))
      logger.warn('DONE!')

   def show(self):
      logger.info('Showing the preference ...')
      q=Preference.objects.all()

      if self.kwargs['name']:
         logger.info('   with the name filter: {0}'.format(self.kwargs['name']))
         if self.kwargs['wildcard'] in self.kwargs['name']:
            q=q.filter(name__icontains=self.kwargs['name'].replace(self.kwargs['wildcard'], ''))
         else:
            q=q.filter(name=self.kwargs['name'])

      if self.kwargs['value']:
         logger.info('   with the value filter: {0}'.format(self.kwargs['value']))
         q=q.filter(value__icontains=self.kwargs['value'])

      if self.kwargs['owner']:
         logger.info('   which belongs to user: {0}'.format(self.kwargs['owner']))
         q=q.filter(owner__username=self.kwargs['owner'])

      if self.kwargs['parent']:
         logger.info('   which belongs to preference: {0}'.format(self.kwargs['parent']))
         q=q.filter(parent__name__iexact=self.kwargs['parent'])
      else:
         q=q.filter(parent__isnull=True)

      for p in q:
         self.output(p)
      logger.warning('There have {0} preference(s) has been shown'.format(len(q)))

   def set(self):
      with transaction.atomic():
         try:
            pref=self.__get_pref__()
            if pref.count()<1: raise Preference.DoesNotExist
            cnt=pref.update(value=self.kwargs['value'])
            logger.info('{0} of Preference(s) has been updated'.format(cnt))
         except Preference.DoesNotExist:
            p=Preference(name=self.kwargs['name'], value=self.kwargs['value'], owner=owner, parent=parent)
            p.save()
            logger.info('The preference<{0}> has been created with value: {1}'.format(p.name, p.value))

   def delete(self):
      pref=self.__get_pref__()
      cnt=pref.count()
      pref.delete()
      logger.warning('{0} of Preference(s) has been deleted'.format(cnt))

   def expRow( self, wr, pref, indent=0 ):
      '''
      Import the specified preference to csv.
      '''
      cnt=0
      tab=self.__getIndent__(indent)
      logger.debug(lm('{0}Exporting preference: {1}::{2}...', tab, pref.id, pref.name))
      wr.writerow([
         pref.name                                    # [0]
         , pref.realValue                             # [1]
         , pref.parent.id if pref.parent else ''      # [2]
         , pref.owner.username if pref.owner else ''  # [3]
         , pref.helptext                              # [4]
         , Preference.TYPES[pref.tipe][1]             # [5]
         , pref.encrypted                             # [6]
         , pref.regex                                 # [7]
      ])
      cnt+=1
      for p in pref.childs:
         cnt+=self.expRow(wr, p, indent+3)
      return cnt

   def expCsv( self ):
      '''
      Import the specified list of preferences to csv.
      '''
      import csv
      f=self.kwargs['file']
      with open(f, 'w', encoding=self.kwargs['encoding']) as fp:
         wr=csv.writer(fp, delimiter=self.kwargs['separator'], quotechar=self.kwargs['quotechar'], quoting=csv.QUOTE_MINIMAL, skipinitialspace=True)
         cnt=0
         for p in self.__get_pref__():
           cnt+=self.expRow(wr, p, 0) 
      logger.info(lm('Exported {0} records', cnt))

   def improw( self, cols, idx=0 ):
      try:
         name=cols[0]
         val=cols[1]
         parent=self.__get_parent__(cols[2])
         owner=self.__get_owner__(cols[3])
         helptext=cols[4]
         tipe=cols[5]
         encrypted=cols[6] in TRUE_VALUES
         regex=cols[7]
         lang=cols[8] if len(cols)>8 else None
         logger.debug('     Importing row: {0}: {1} [{2}]'.format(idx, name, 'encrypted' if encrypted else 'clear-text'))

         self.kwargs['name']=name
         pref=self.__get_pref__(name=name, owner=owner, parent=parent, lang=lang)
         if pref.count()<1: raise Preference.DoesNotExist
         for p in pref:
            p.encrypted=encrypted
            p.helptext=helptext
            p.tipe=tipe
            p.regex=regex
            #The value must be the last steps to set due to validation. Otherwise, once importing/assign a new value into this field, the last validation rule may be applied incorrectly
            p.value=val 
            p.save()
      except Preference.DoesNotExist:
         Preference(name=name, _value=val, owner=owner, parent=parent, encrypted=encrypted, helptext=helptext, regex=regex, lang=lang).save()
      except:
         logger.debug(cols)
         logger.exception('Error when handling the column')
         raise

   def impXlsx( self, f ):
      '''
      Import xlsx file.
      '''
      from openpyxl import load_workbook
      wb=load_workbook(filename=f)
      ws=wb.active
      logger.info('   Importing worksheet: {0}!{1}'.format(f, ws.title))
      cnt=0
      with transaction.atomic():
         for r in range(1, ws.max_row+1):
            cols=list()
            name=ws.cell(row=r, column=1).value
            if isinstance(name, str): name=name.strip()
            if not name: continue #Skip the row when it has no pref.name
            if r==1 and (name.upper()=='ID' or name.upper()=='NAME' or name.upper()=='ID/Name'): continue #Skip the first row if header row
            cols.append(name) #Name/ID
            cols.append(ws.cell(row=r, column=2).value) #Value
            cols.append(ws.cell(row=r, column=3).value) #Parent
            cols.append(ws.cell(row=r, column=4).value) #Owner
            cols.append(ws.cell(row=r, column=5).value) #Reserved
            cols.append(ws.cell(row=r, column=6).value) #Tipe
            cols.append(ws.cell(row=r, column=7).value) #encrypted
            self.improw( cols, r )
            cnt+=1
      logger.info('   Imported {0} row(s)'.format(cnt))

   def impCsv( self, f ):
      '''
      Import the csv file.
      '''
      import csv
      with transaction.atomic():
         logger.info('   Importing csv: {0}'.format(f))
         cnt=0
         with open(f, 'r', encoding=self.kwargs['encoding']) as fp:
            if self.kwargs['quotechar']:
               rows=csv.reader(fp, delimiter=self.kwargs['separator'], quotechar=self.kwargs['quotechar'], quoting=csv.QUOTE_MINIMAL, skipinitialspace=True)
            else:
               rows=csv.reader(fp, delimiter=self.kwargs['separator'], quoting=csv.QUOTE_NONE, skipinitialspace=True)
            for row in rows:
               if len(row)<1: continue #Skip the empty row
               name=row[0].strip()
               if not name: continue #Skip the row when it has no name
               if cnt==0 and (name.upper()=='ID' or name.upper()=='NAME' or name.upper()=='ID/NAME'): continue #Skip the first row if header row
               self.improw( row, cnt )
               cnt+=1
         logger.info('   Imported {0} row(s)'.format(cnt))

   def impdir( self, d ):
      if os.path.isdir(d):
         logger.info('Importing directory: {0}'.format(d))
      else:
         logger.warning('This is not the directory: {0}'.format(d))
         return

      cnt=0
      with transaction.atomic():
         p=Preference.objects.pref('IMPORTED_PREFERENCES', returnValue=False)
         p.helptext='<p>Sysetm use only! <strong>DO NOT MODIFY</strong> youself unless you understand the risk.</p>'
         p.save()
         for f in os.listdir(d):
            if not (f.upper().endswith('.XLSX') or f.upper().endswith('.CSV')): continue #only support *.xlsx and *.csv
            f=os.path.join(d, f)
            try:
               Preference.objects.get(name=f, parent=p)
               if self.kwargs['force']: raise Preference.DoesNotExist
            except Preference.DoesNotExist:
               self.impfile( f )
               cnt+=1
               Preference(name=f, parent=p).save()
      logger.debug('Imported {0} file(s)'.format(cnt))

   def impfile( self, f ):
      if not (os.path.isfile(f) and os.access(f, os.R_OK)):
         logger.warning('The file is not readable: {0}'.format(f))
         return
      fn=f.lower()
      if fn.endswith('.xlsx'):
         self.impXlsx(f)
      elif fn.endswith('.csv'):
         self.impCsv(f)
      else:
         logger.info('Unsupported file: {0}'.format(f))

   def imppath( self, p, parent=None):
      name, seq=self.__get_name__(os.path.basename(p))
      if os.path.isdir(p):
         try:
            pref=self.__get_pref__(name=name)
            if pref.count()<1: raise Preference.DoesNotExist
            pref=pref[0]
         except Preference.DoesNotExist:
            pref=Preference(name=name, parent=parent)
         pref.tipe=AbstractPreference.TYPE_FILEPATH
         pref.sequence=seq
         pref.save()

         for f in os.listdir(p):
            path=os.path.join(p, f)
            self.imppath(path, pref)

         #Handling the ordering after import all the childs
         ord=1
         for c in pref.childs:
            c.sequence=ord
            c.save()
            ord+=1
      else:
         try:
            pref=self.__get_pref__(name=name)
            if pref.count()<1: raise Preference.DoesNotExist
            pref=pref[0]
         except Preference.DoesNotExist:
            pref=Preference(name=name, parent=parent)
         pref.pathValue=p if os.path.isabs(p) else os.path.abspath(p)
         pref.tipe=AbstractPreference.TYPE_FILEPATH
         pref.sequence=seq
         pref.save()

   def imp(self):
      disableOrder=getattr(settings, 'DISABLE_REORDER', False)
      setattr(settings, 'DISABLE_REORDER', True) #Disable the re-ordering features during importing
      try:
         f=self.kwargs['file']
         if self.kwargs['filepath']:
            self.imppath(f)
         elif os.path.isdir(f):
            self.impdir(f)
         elif os.path.isfile(f):
            self.impfile(f)
      finally:
         setattr(settings, 'DISABLE_REORDER', disableOrder) #Resume the re-ordering features after importing

   def gensecret(self):
      from webframe.models import AbstractPreference
      key=AbstractPreference.__getSecret__()
      logger.warning(lm('Your secret is: {0}', key))

   def gendoc(self):
      from django.shortcuts import render
      from django.template import loader, Template, Context
      from webframe.providers import template_injection, fmt_injection
      tmpl=getattr(self.kwargs, 'tmpl', 'webframe/prefDoc.html') 
      logger.warning(lm('Generating the documents according template: {0}', tmpl))
      tmpl=loader.get_template(tmpl)

      params=dict()
      params.update(template_injection(None))
      params.update(fmt_injection(None))
      #params['target']=Preference.objects.filter(parent__isnull=True)
      params['target']=self.__get_pref__()
      params['TYPES']=Preference.TYPES
      params['now']=getTime('now')
      txt=tmpl.render(params)
      output=self.kwargs.get('file')
      if not output: output='prefsDoc.html'
      logger.warning(lm('Generated! Outputing into: {0}', output))
      with open(output, 'w') as f:
         f.write(txt)
